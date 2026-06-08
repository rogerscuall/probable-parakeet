#!/usr/bin/env python3
"""Generate the arista_evpn_vxlan XLSX template from the label YAML.

The structure (sheets, columns, required flags, enum options) is derived
directly from the `content.validation` blocks in arista_evpn_vxlan.yaml so the
template stays authoritative with the config the app validates against.
"""
from __future__ import annotations

import sys
from collections import OrderedDict

import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

YAML_PATH = sys.argv[1]
OUT_PATH = sys.argv[2]

HEADER_FILL = PatternFill(start_color="FFDDDDDD", end_color="FFDDDDDD", fill_type="solid")
HEADER_FONT = Font(bold=True)


def collect_sheets(cfg: dict) -> "OrderedDict[str, list[dict]]":
    """Return {sheet_name: [ {column, field_type, required, options}, ... ]}.

    Columns are collected in the order they appear in the YAML validation,
    de-duplicated per sheet.
    """
    sheets: "OrderedDict[str, list[dict]]" = OrderedDict()

    # Seed sheet order from each label's xlsx_sheets so empty/validation-less
    # sheets still appear in a stable order.
    for label in cfg.get("labels", []):
        for sheet in label.get("content", {}).get("xlsx_sheets", []) or []:
            sheets.setdefault(sheet, [])

    for label in cfg.get("labels", []):
        for rule in label.get("content", {}).get("validation", []) or []:
            sheet = rule["sheet"]
            cols = sheets.setdefault(sheet, [])
            if any(c["column"] == rule["column"] for c in cols):
                continue
            cols.append(
                {
                    "column": rule["column"],
                    "field_type": rule.get("field_type", ""),
                    "required": bool(rule.get("required", False)),
                    "options": rule.get("options") or [],
                }
            )
    return sheets


def build_instructions(ws, design_name: str) -> None:
    lines = [
        f"Welcome to the {design_name} Configuration Template",
        "",
        "Instructions:",
        "1. Fill out the sheets that correspond to the features you selected.",
        "2. Follow the column headers exactly - do not rename or reorder them.",
        "3. Do not modify or delete the header row.",
        "4. Required columns must be populated for every row.",
        "5. Columns with a dropdown only accept the listed values.",
        "6. Refer to the generated design document for field-by-field guidance.",
    ]
    for line in lines:
        ws.append([line])
    ws["A1"].font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 80


def add_sheet(wb: Workbook, name: str, columns: list[dict]) -> None:
    ws = wb.create_sheet(title=name)
    if not columns:
        return
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col["column"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

        # Header comment documents type / required / allowed values.
        note = f"type: {col['field_type']}\nrequired: {'yes' if col['required'] else 'no'}"
        if col["options"]:
            note += "\nallowed: " + ", ".join(col["options"])
        cell.comment = Comment(note, "template-generator")

        # Width: a touch wider than the header for readability.
        ws.column_dimensions[cell.column_letter].width = max(16, len(col["column"]) + 4)

        # Dropdown for enum columns.
        if col["options"]:
            formula = '"' + ",".join(col["options"]) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=not col["required"])
            ws.add_data_validation(dv)
            dv.add(f"{cell.column_letter}2:{cell.column_letter}1000")

    ws.freeze_panes = "A2"


def main() -> None:
    with open(YAML_PATH) as fh:
        cfg = yaml.safe_load(fh)

    design_name = cfg.get("name", "Arista EVPN VXLAN")
    sheets = collect_sheets(cfg)

    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    build_instructions(instructions, design_name)

    for name, columns in sheets.items():
        add_sheet(wb, name, columns)

    wb.save(OUT_PATH)

    print(f"Wrote {OUT_PATH}")
    print(f"Sheets ({len(sheets) + 1}): Instructions, " + ", ".join(sheets))
    for name, columns in sheets.items():
        cols = ", ".join(c["column"] for c in columns)
        print(f"  - {name}: {cols}")


if __name__ == "__main__":
    main()
